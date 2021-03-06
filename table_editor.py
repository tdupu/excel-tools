from openpyxl import *
from table_functions import *

def is_subdictionary(A,B):
    return set(A.items()).issubset(B.items())

class SheetObject:
    """
    A wrapper to help work with excel_spreadsheets for sheets of openpyxl.
    
    Assumptions:
    --the spreadsheed has the first column a list of strings
    (these will be the keys for our dictionaries)
    --we never want to add more keys to our sheet.
    
    > from openpyxl import *
    >
    
    """
    
    
    def __init__(self,user_filename,user_sheetname):
        """
        To instantiate such a class we just need to pass it an openpyxl sheet.
        """
        #read off the first row of the spreadsheet keep track of which column has which entry
        self.sheetname = user_sheetname
        self.filename = user_filename
        self._workbook = load_workbook(self.filename)
        self._sheet = self._workbook[self.sheetname]
        self.keys = []
        self.column_dict = {}

        j=0
        for value in self._sheet.iter_cols(min_row=1, max_row=1,values_only=True):
            self.keys.append(value[0])
            self.column_dict[j+1] = self.keys[j]
            j=j+1
            
        self.number_of_keys = len(self.keys)
        self.set_of_keys = set(self.keys)
        
    #def __iter__(self):
    #    return self
        
    #def __next__(self):
    #    not implemented
    
    def is_valid_entry(self, new_entry, is_full=False):
        """
        New entries are assumed to be dictionaries.
        """
        set_of_entry_keys = set(list(new_entry.keys()))
        
        if (is_full == False):
            return set_of_entry_keys.issubset(self.set_of_keys)
            
        elif (is_full == True):
            return set_of_entry_keys == self.set_of_keys
            
        else:
            raise ValueError('is_full must be True or False')

        
    def append(self, new_entry):
        """
        Takes a dictionary input and if its keys match the keys for the spreadsheet it will make a new row.
        """
        if self.is_valid_entry(new_entry):
            if not self.has_entry(new_entry):
                """
                print("column dict: %s \n <br>" % self.column_dict)
                """
                new_row = [new_entry[self.column_dict[i+1]] for i in range(self.number_of_keys)]
                self._sheet.append(new_row)
                #self.save() #appears after this function, I don't remember if this matters.
            
        else:
            A = set(new_entry.keys())
            B = self.set_of_keys
            C = A.symmetric_difference(B)
            message = "symmetric difference of keys: \n"
            for c in C:
                message = message + ("%s \n" % c)

            raise ValueError("Append failed. Entry keys do not match .xlsx keys. \n" + message)
            
    
    
    
    def get_with_index(self,partial_entry):
        """
        --INPUT: partial_entry is a dictionary which has a subset of self.keys for entries.
        --OUTPUT: this function will return all the elements of the spreadsheet as dictionaries whose entries matches those that we searched for.
    
        (If you pass an empty dict, it should give you all the entries as a dictionary)
        """
        a = partial_entry
        X = self.is_valid_entry(a)
        
        if X:
        #Y=True
        #if Y:
            entry_keys = partial_entry.keys()
            
            matches = []
            indices = []
            
            i=1
            for row in self._sheet.iter_rows(min_row=2,values_only=True):
                i=i+1
                #first convert the row to a dictionary
                row_as_dictionary = self.row_as_list_to_dict(row)
                
                #if the row is a match, throw the dictionary into the list of matches
                if is_subdictionary(partial_entry,row_as_dictionary):
                    matches.append(row_as_dictionary)
                    indices.append(i)
            
            #print([matches,indices])
            return [matches,indices]
            
        else:
            raise ValueError(f'\n {partial_entry} is not a valid entry')
            
    def get(self,entry):
        return self.get_with_index(entry)[0]
            
    def get_all(self):
        """
        return all entries.
        """
        return self.get({})
            
    def get_index(self,entries):
        """
        get a list of indices for a given set of entries.
        entries is a list of dictionaries.
        all dictionaries must match all the keys.
        """
        
        entries_with_index = self.get_with_index(entries)
        #print('getting entries with index \n')
        #print(entries_with_index)
        return entries_with_index[1]
        
        
        
        raise NotImplementedError("Didn't get to this yet")
        
    def get_by_excel_row_index(self,x):
        """
        Get the xth row and return it as a list.
        """
        row = {}
        j=0
        for value in self._sheet.iter_cols(min_row=x, max_row=x,values_only=True):
            row[self.keys[j]]=value[0]
            j=j+1
        return row
        
    def replace(self,old_entry,new_entry, is_subentry=False):
        if is_subentry==True:
            keep_going = is_subdictionary(old_entry,new_entry)
            if not keep_going:
                raise ValueError("The new entry does not extend the old entry")
        
        search_result = [entry for entry in self.get({'submission_number':old_entry['submission_number']})]
        n=len(search_result)
        if n==1:
            self.remove(entries=search_result)
            self.append(new_entry)
            return "The entry has been updated. Remember to save!"
        if n==0:
            raise ValueError("Cannot replace. No such entry.")
        if n>0:
            raise ValueError("Cannot replace. There is not a unique entry for this query.")
        
    
    
    def remove(self, entries=[], list_of_row_indices=[]):
        """
        Take a list of dictionaries or a list of rows and remove entries.
        """
        n = len(list_of_row_indices)
        m = len(entries)
        
        if (m!=0) and (n!=0):
            raise ValueError("input must be a list of dictionaries or a list of row indices but not both")
        
        elif n!=0 and m==0:
            #print("removing the following: %s" % list_of_row_indices)
            #print(list_of_row_indices)
            
            sorted_indices_for_deletion = sorted(list_of_row_indices, reverse=True) #remove large to small indices so
            #if len(sorted_indices_for_deletion)==0:
            #    return "no rows were deleted."
            #else:
            for i in sorted_indices_for_deletion:
                self._sheet.delete_rows(idx=i)
                
            #print('all: %s' % self.get_all())
            return "rows have been removed"
                
        elif n==0 and m!=0:
            #print(m)
            #print(entries)
            
            list_of_row_ind = []
            
            for entry in entries:
                
                if self.is_valid_entry(entry,is_full=True)==False:
                    raise ValueError("all entries must specify a complete set of keys")
                
                
                myindex = self.get_index(entry)[0]
                #print('index: %s ' % myindex)
                list_of_row_ind.append(myindex)
                
            #print('list: %s' % list_of_row_ind)
            self.remove(list_of_row_indices=list_of_row_ind)
            
    def number_of_entries(self):
        return self._sheet.max_row-1
        
    def row_as_list_to_dict(self, row):
        """
        Converts a given row to a dictionary.
        The row needs to be a list or tuple with values only!
        """
        row_as_dictionary = {}
        for i in range(self.number_of_keys):
            row_as_dictionary[self.keys[i]] = row[i]
            
        return row_as_dictionary
        
    def save(self,save_name=None):
        """
        Saves the excel file.
        save_name must be a string
        """
        if save_name == None:
            self._workbook.save(filename=self.filename)
        else:
            self._workbook.save(filename=save_name)
            
    def has_entry(self,mydict):
        """
        Given a dictionary mydict, search the spreadsheet to see if this exact same entry already exists. ON THE NOSE!
        """
        if self.is_valid_entry(mydict):
            matching_entries = self.get(mydict)
            found_exact_same_entry = False
            for entry in matching_entries:
                if entry == mydict:
                    found_exact_same_entry = True
                    break
            return found_exact_same_entry
            
        else:
            raise ValueError("dictionary keys do not match column headings in the spreadsheet")
            
    def set_entry(self,sub,key,entry):
        newsub=copyd(sub)
        newsub[key]=entry
        self.replace(sub,newsub)
        return "entry for '%s': %s -> %s" % (key,sub[key],entry)
         
